import re
import json
import pandas as pd
import os
import math
from fuzzywuzzy import fuzz
import warnings
from datetime import datetime
import openpyxl
from typing import Dict, List


# Suppress openpyxl warning about missing default styles
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# Load the mapping file as a dictionary (assuming it's stored in 'mapping.json')
with open('column_mapping.json', 'r') as file:
    mapping_data = json.load(file)


def extract_iso_and_date(filename, similarity_threshold=80):
    # Match patterns like "FirstDirect Financial May2024.xls", "RAC Sept 2024 residuals"
    # pattern = r'^(.*?)(?:\s+)?([A-Za-z]+)\s?(\d{4})(?:\s.*)?$'
    pattern = r'^(.*?)(?:[_\s]+)?([A-Za-z]+)[_\s]?(\d{4})(?:\s.*)?$'
    match = re.match(pattern, filename.strip())

    if match:
        extracted_name = match.group(1).strip()  # Extracted name for ISO lookup
        month = match.group(2)                  # Extracted month
        year = match.group(3)                   # Extracted year

        # Normalize the extracted name (fix spacing issues)
        extracted_name_normalized = ' '.join(extracted_name.split())

        # Perform fuzzy matching to find the best match
        best_match = None
        highest_similarity = 0
        
        for key in mapping_data.keys():
            similarity = fuzz.ratio(key.lower(), extracted_name_normalized.lower())
            if similarity > highest_similarity and similarity >= similarity_threshold:
                best_match = key
                highest_similarity = similarity

        if not best_match:
            return {"error": f"No matching ISO found for '{extracted_name}' (similarity below {similarity_threshold}%)"}

        # Retrieve the ISO from the best match
        iso = mapping_data[best_match].get('iso')

        # Convert month to a valid number
        month_number = {
            'January': '01', 'February': '02', 'March': '03', 'April': '04',
            'May': '05', 'June': '06', 'July': '07', 'August': '08',
            'September': '09', 'October': '10', 'November': '11', 'December': '12',
            'Jan': '01', 'Feb': '02', 'Mar': '03', 'Apr': '04', 'Jun': '06',
            'Jul': '07', 'Aug': '08', 'Sep': '09', 'Oct': '10', 'Nov': '11', 'Dec': '12', 'Sept': '09'
        }.get(month.capitalize())

        if month_number:
            formatted_date = f"{year}-{month_number}-01"
            return {"iso": iso, "date": formatted_date}
        else:
            return {"error": "Invalid month in filename"}
    else:
        return {"error": "Filename format is invalid"}
    

def load_mapping(mapping_file):
    with open(mapping_file, 'r') as f:
        return json.load(f)

def is_valid_merchant_id(mid):
    return pd.notna(mid) and (isinstance(mid, (int, float)) or (isinstance(mid, str) and mid.replace('-', '').isdigit()))

def find_data_rows(df, mapping):
    data_rows = []
    header_row = None
    mapped_columns = [col for col in mapping.values() if col]

    for i, row in df.iterrows():
        if all(col in row.values for col in mapped_columns):
            header_row = i
            break

    if header_row is not None:
        df.columns = df.iloc[header_row]
        print(df.iloc[header_row])
        for i, row in df.iloc[header_row + 1:].iterrows():
            try:
                if is_valid_merchant_id(row[mapping['mid']]):
                    data_rows.append(i)
            except KeyError as e:
                print(f"KeyError in find_data_rows: {e}")
                return None, []

    return header_row, data_rows

def ceil_to_two_decimals(value):
    if pd.notna(value):  
        try:
            float_value = float(value)
            return math.ceil(float_value * 100) / 100  
        except (ValueError, TypeError):
            return value
    return value

def normalize_name(name):
    """Remove special characters and normalize the string for easier matching."""
    return re.sub(r'[^a-zA-Z0-9]', '', name).lower()

def fuzzy_match_filename(filename, mappings):
    filename_base = normalize_name(os.path.splitext(filename)[0])
    for mapping_key in mappings.keys():
        mapping_base = normalize_name(mapping_key)
        if mapping_base in filename_base:  # Check if the key is a substring of the filename
            return mapping_key
    return None

def process_file(file_path, mapping, iso_name):
    try:
        if file_path.filename.endswith('.csv'):
            df = pd.read_csv(file_path, header=None)
        else:
            df = pd.read_excel(file_path, header=None)
    except Exception as e:
        print(f"Error reading file {file_path}: {str(e)}")
        return None

    print(f"\nProcessing file: {file_path}")
    header_row, data_rows = find_data_rows(df, mapping)

    if header_row is not None and data_rows:
        df.columns = df.iloc[header_row]
        df_filtered = df.iloc[data_rows].copy()

        # Filter out empty string mappings and select only available columns
        valid_columns = {k: v for k, v in mapping.items() if v}  # Ignore empty string columns
        try:
            df_final = df_filtered[[column for column in valid_columns.values()]]
        except KeyError as e:
            print(f"KeyError when selecting columns: {e}")
            return None

        df_final.columns = valid_columns.keys()

        string_columns = ['mid', 'dba']
        for col in string_columns:
            if col in df_final.columns:
                df_final.loc[:, col] = df_final[col].astype(str)

        numerical_columns = ['volume', 'total_residual', 'paydiverse_residual']

        # Ensure df_final is a standalone copy to avoid SettingWithCopyWarning
        df_final = df_final.copy()

        # Preprocess and set missing numerical columns to 0.00 if they don't exist
        for col in numerical_columns:
            if col in df_final.columns:
                # Remove any '$' or ',' characters and convert to float
                df_final[col] = df_final[col].astype(str).replace({r'\$': '', ',': ''}, regex=True)
                df_final.loc[:, col] = pd.to_numeric(df_final[col], errors='coerce')
                df_final.loc[:, col] = df_final[col].apply(ceil_to_two_decimals).fillna(0.00)  # Fill NaN with 0.00
            else:
                df_final.loc[:, col] = 0.00  # Set missing columns to 0.00 if they don't exist
    
        df_final = df_final.copy()  # Avoid SettingWithCopyWarning
        df_final.loc[:, 'iso'] = iso_name  # Add iso column explicitly

        return df_final.reset_index(drop=True)
    else:
        print(f"Could not find header or data rows.")

    return None

# Updated function to take a file as a parameter and return JSON
def process_uploaded_file(file, mapping_file):
    mappings = load_mapping(mapping_file)
    file_type = fuzzy_match_filename(file.filename, mappings)

    if not file_type:
        return {"error": "No mapping found for file"}

    iso_name = mappings[file_type].get('iso', '')
    mapping_data = mappings[file_type]['mapping']

    extracted_data = process_file(file, mapping_data, iso_name)

    if extracted_data is None:
        return {"error": "Data extraction failed"}
    else:
        return extracted_data


def process_authorize_file(file_path):
    """
    Processes an Excel file and extracts specific columns to compute paydiverse_residual.

    Args:
        file_path (str): The path to the Excel file to be processed.

    Returns:
        pd.DataFrame: The processed DataFrame with selected columns.
    """
    try:
        # Load the Excel file
        excel_data = pd.ExcelFile(file_path)

        # Load data from the first sheet (regardless of its name)
        sheet_data = excel_data.parse(excel_data.sheet_names[0])

        # Select the required columns and create the new calculated column
        selected_data = sheet_data[['MID', 'Reference ID', 'Collection Attempts', 'Collection Returns/Subsidized Fees']].copy()
        selected_data.rename(columns={'Reference ID': 'dba', 'MID': 'mid'}, inplace=True)

        # Calculate paydiverse_residual
        selected_data['paydiverse_residual'] = (
            selected_data['Collection Attempts'] - abs(selected_data['Collection Returns/Subsidized Fees'])
        )

        # Add new columns Volume and Total_residual
        selected_data['volume'] = 0.00
        selected_data['total_residual'] = 0.00

        # Drop intermediate calculation columns
        selected_data = selected_data[['mid', 'dba', 'volume', 'total_residual', 'paydiverse_residual']]

        # Combine rows with the same 'mid' by summing 'paydiverse_residual'
        grouped_data = (
            selected_data.groupby('mid', as_index=False)  # Group by 'mid'
            .agg({
                'dba': 'first',  # Keep the first 'dba'
                'volume': 'first',  # Keep the first 'volume' (same value for duplicates)
                'total_residual': 'first',  # Keep the first 'total_residual' (same value for duplicates)
                'paydiverse_residual': 'sum'  # Sum the 'paydiverse_residual'
            })
        )

        # Final processed data
        selected_data = grouped_data
        return selected_data
    except Exception as e:
        print(f"Error processing file: {e}")


def process_nuvei_file(file_path, sheet_name = 'PPI'):
    # Load the sheet
    try:
        # Load the sheet data
        sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, header=0)

        # Clean column names by stripping whitespace and line breaks
        sheet_data.columns = sheet_data.columns.str.replace('\n', ' ').str.strip()

        # Define columns to extract and their new names
        columns_to_extract = {
            'MID': 'mid',
            'DBA Name': 'dba',
            'Total Profit': 'total_residual',
            'Agent Net': 'paydiverse_residual'
        }

        # Check if all required columns exist
        missing_columns = [col for col in columns_to_extract.keys() if col not in sheet_data.columns]
        if missing_columns:
            raise ValueError(f"The following required columns are missing: {missing_columns}")

        # Ensure volume-related columns exist
        volume_columns = ['V/MC/Discover Vol', 'AMEX Vol']
        missing_volume_columns = [col for col in volume_columns if col not in sheet_data.columns]
        if missing_volume_columns:
            raise ValueError(f"The following volume-related columns are missing: {missing_volume_columns}")

        # Extract and rename the required columns
        extracted_data = sheet_data[list(columns_to_extract.keys())].rename(columns=columns_to_extract)

        # Add the volume column as the sum of 'V/MC/Discover Vol' and 'AMEX Vol'
        sheet_data['volume'] = sheet_data['V/MC/Discover Vol'].fillna(0) + sheet_data['AMEX Vol'].fillna(0)

        # Add the new 'volume' column to the extracted data
        extracted_data['volume'] = sheet_data['volume']

        # Filter rows where 'mid' is an integer
        extracted_data = extracted_data[pd.to_numeric(extracted_data['mid'], errors='coerce').notna()]

        # Convert 'mid' to int for consistency
        extracted_data['mid'] = extracted_data['mid'].astype(int)

        # Reorder the columns
        desired_order = ['mid', 'dba', 'volume', 'total_residual', 'paydiverse_residual']
        extracted_data = extracted_data[desired_order]

        return extracted_data

    except Exception as e:
        print(f"Error: {e}")


def process_nmi_file(file_path):
    # Load the sheet dynamically since there is only one sheet
    try:
        # Load the data
        sheet_data = pd.read_excel(file_path, sheet_name=0, header=0)

        # Clean column names
        sheet_data.columns = sheet_data.columns.str.replace('\n', ' ').str.strip()

        # Define columns to extract and their new names
        columns_to_extract = {
            'MID': 'mid',
            'Merchant': 'dba',
            'Total Commission': 'paydiverse_residual',
            'Total Resale Price': 'total_residual'
        }

        # Check if all required columns exist
        missing_columns = [col for col in columns_to_extract.keys() if col not in sheet_data.columns]
        if missing_columns:
            raise ValueError(f"The following required columns are missing: {missing_columns}")

        # Extract and rename the required columns
        extracted_data = sheet_data[list(columns_to_extract.keys())].rename(columns=columns_to_extract)

        # Add missing 'volume' column with default value 0.00
        extracted_data['volume'] = 0.00

        # Group by 'mid' and aggregate
        aggregated_data = (
            extracted_data.groupby(['mid', 'dba'], as_index=False)
            .agg({'volume': 'sum', 'total_residual': 'sum', 'paydiverse_residual': 'sum'})
        )

        # Reorder columns as per the specified order
        final_columns = ['mid', 'dba', 'volume', 'total_residual', 'paydiverse_residual']
        aggregated_data = aggregated_data[final_columns]

        return aggregated_data

    except Exception as e:
        print(f"Error: {e}")


def process_cwa_file(file_path, sheet_name = 'Residual Payments - Schedule A'):
    # Load the sheet
    sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # Locate the header row by checking for 'ACCOUNT'
    header_row = None
    for i, row in sheet_data.iterrows():
        if row.astype(str).str.contains('ACCOUNT', case=False, na=False).any():
            header_row = i
            break

    if header_row is None:
        raise ValueError("Header row containing 'ACCOUNT' not found.")

    # Reload the data with the correct header
    sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)

    # Clean column names by stripping whitespace and line breaks
    sheet_data.columns = sheet_data.columns.str.replace('\n', ' ').str.strip()

    # Define columns to extract and their new names
    columns_to_extract = {
        'ACCOUNT': 'mid',
        'DBA Name': 'dba',
        'Total Sales': 'volume',
        'Net Profit': 'total_residual',
        'Payment to Sales Agent': 'paydiverse_residual'
    }

    # Check if all required columns exist
    missing_columns = [col for col in columns_to_extract.keys() if col not in sheet_data.columns]
    if missing_columns:
        raise ValueError(f"The following required columns are missing: {missing_columns}")

    # Extract and rename the required columns
    extracted_data = sheet_data[list(columns_to_extract.keys())].rename(columns=columns_to_extract)

    # Filter rows where 'mid' is an integer
    extracted_data = extracted_data[pd.to_numeric(extracted_data['mid'], errors='coerce').notna()]

    # Convert 'mid' to int for consistency
    extracted_data['mid'] = extracted_data['mid'].astype(int)

    # Output the resulting data
    return extracted_data


def process_rac_file(file_path):
    # Dynamically select the single sheet
    excel_data = pd.ExcelFile(file_path)
    sheet_name = excel_data.sheet_names[0]

    # Load the sheet
    sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, header=0)

    # Clean column names by stripping whitespace and line breaks
    sheet_data.columns = sheet_data.columns.str.replace('\n', ' ').str.strip()

    # Define columns to extract and their new names
    columns_to_extract = {
        'MID': 'mid',
        'MERCHANT': 'dba',
        'NET SALES': 'volume',
        'SPLIT TO RAC - 70%': 'total_residual',
        'SPLIT TO AGENT - 70%': 'paydiverse_residual'
    }

    # Check if all required columns exist
    missing_columns = [col for col in columns_to_extract.keys() if col not in sheet_data.columns]
    if missing_columns:
        raise ValueError(f"The following required columns are missing: {missing_columns}")

    # Extract and rename the required columns
    extracted_data = sheet_data[list(columns_to_extract.keys())].rename(columns=columns_to_extract)

    # Filter rows where 'mid' is an integer
    extracted_data = extracted_data[pd.to_numeric(extracted_data['mid'], errors='coerce').notna()]

    # Convert 'mid' to int for consistency
    extracted_data['mid'] = extracted_data['mid'].astype(int)

    return extracted_data


 # Define mappings for each sheet
sheet_mappings = {
    'Digipay': {
        'mid': {'MID', 'Merchant ID', 'Merchant Identifier'},
        'dba': {'DBA', 'Merchant', 'DBAName'},
        'volume': {'Volume', 'BC Volume', 'volume', 'Total SalesCount'},
        'total_residual': {'Revenue', 'Res. Earned', 'Payee Income', 'Total', 'Payout'},
        'paydiverse_residual': {'Partner Comm.', 'Partner Comm'}
    },
    'EMS': {
        'mid': {'MID', 'Merchant ID', 'Merchant Identifier'},
        'dba': {'DBA', 'Merchant', 'DBAName'},
        'volume': {'Volume', 'BC Volume', 'volume'},
        'total_residual': {'Revenue', 'Res. Earned', 'Payee Income', 'Total'},
        'paydiverse_residual': {'Partner Comm.', 'Partner Comm'}
    },
    'EMS_OR': {
        'mid': {'MID', 'Merchant ID', 'Merchant Identifier'},
        'dba': {'DBA', 'Merchant', 'DBAName'},
        'volume': {'Volume', 'BC Volume', 'volume'},
        'total_residual': {'Revenue', 'Res. Earned', 'Payee Income', 'Total'},
        'paydiverse_residual': {'Partner Comm.', 'Partner Comm', 'Partner %2'}
    },
    'First_Data': {
        'mid': {'MID', 'Merchant ID', 'Merchant Identifier'},
        'dba': {'DBA', 'Merchant', 'DBAName'},
        'volume': {'Volume', 'BC Volume', 'volume'},
        'total_residual': {'Revenue', 'Res. Earned', 'Payee Income', 'Total'},
        'paydiverse_residual': {'Partner Comm.', 'Partner Comm'}
    },
    'NXGEN_Elavon': {
        'mid': {'MID', 'Merchant ID', 'Merchant Identifier'},
        'dba': {'DBA', 'Merchant', 'DBAName'},
        'volume': {'Volume', 'BC Volume', 'volume'},
        'total_residual': {'Revenue', 'Res. Earned', 'Payee Income', 'Total'},
        'paydiverse_residual': {'Partner Comm.', 'Partner Comm'}
    },
    'TMS': {
        'mid': {'MID', 'Merchant ID', 'Merchant Identifier'},
        'dba': {'DBA', 'Merchant', 'DBAName'},
        'volume': {'Volume', 'BC Volume', 'volume'},
        'total_residual': {'Revenue', 'Res. Earned', 'Payee Income', 'Total', 'Total Residuals Earned'},
        'paydiverse_residual': {'Partner Comm.', 'Partner Comm'}
    },
    'CMS_Nexio': {
        'mid': {'MID', 'Merchant ID', 'Merchant Identifier'},
        'dba': {'DBA', 'Merchant', 'DBAName', 'Merchant Name'},
        'volume': {'Volume', 'BC Volume', 'volume'},
        'total_residual': {'Revenue', 'Res. Earned', 'Payee Income', 'Total', 'Agent Residual'},
        'paydiverse_residual': {'Partner Comm.', 'Partner Comm'}
    },
    'CMS': {
        'mid': {'MID', 'Merchant ID', 'Merchant Identifier'},
        'dba': {'DBA', 'Merchant', 'DBAName', 'Merchant Name'},
        'volume': {'Volume', 'BC Volume', 'volume'},
        'total_residual': {'Revenue', 'Res. Earned', 'Payee Income', 'Total', 'Agent Residual'},
        'paydiverse_residual': {'Partner Comm.', 'Partner Comm'}
    }
}

# Define ISO name mappings
iso_name_mappings = {
    'First_Data': 'FD',
    'NXGEN_Elavon': 'NXGEN',
    'EMS_OR': 'EMS OR',
    'CMS_Nexio': 'CMS',
    'CMS': 'CMS'
}

def process_paymentcloud_file(file_path, sheet_mappings = sheet_mappings, iso_name_mappings = iso_name_mappings):
    excel_file = pd.ExcelFile(file_path)
    
    # List to store all DataFrames
    all_data = []
    
    # Process each sheet that exists in the file and has defined mappings
    for sheet_name in excel_file.sheet_names:
        if sheet_name in sheet_mappings:
            column_mapping = sheet_mappings[sheet_name]
            # Read the current sheet into a DataFrame
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Create a dictionary to hold the extracted columns
            extracted_columns = {}
            
            for target_col, possible_names in column_mapping.items():
                # Find the first matching column name in the current sheet
                for col in possible_names:
                    if col in df.columns:
                        extracted_columns[target_col] = df[col]
                        break  # Stop checking after finding the first match
                
                # If no matching column is found, add an empty column for consistency
                if target_col not in extracted_columns:
                    extracted_columns[target_col] = pd.Series(dtype='object')

            # Create a new DataFrame with the extracted columns
            temp_df = pd.DataFrame(extracted_columns)

            # Ensure the MID is treated as an integer and convert it to a string (to prevent scientific notation)
            if 'mid' in temp_df.columns:
                # Filter rows where MID is a valid integer (not NaN, and numeric)
                temp_df = temp_df[temp_df['mid'].apply(lambda x: pd.notna(x) and isinstance(x, (int, float)) and x == int(x))]
                # Convert MID to an integer and then to a string to avoid scientific notation
                temp_df['mid'] = temp_df['mid'].astype(int).astype(str)

            # Determine the ISO name using the mapping
            adjusted_iso_name = iso_name_mappings.get(sheet_name, sheet_name)
            temp_df.insert(0, 'iso', f"Payment Cloud {adjusted_iso_name}")

            # Append the processed DataFrame to the list
            all_data.append(temp_df)

    # Combine all DataFrames into one
    combined_df = pd.concat(all_data, ignore_index=True)

    return combined_df


def process_highrisk_file(file_path):
    # Load the Excel file
    excel_file = pd.ExcelFile(file_path)

    # Initialize a list to store MID, DBA, volume, and paydiverse_residual information
    mids_data = []

    # Loop through all sheets
    for sheet_name in excel_file.sheet_names:
        sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Search for "Mid#" in the sheet
        mid_cell = sheet_data[sheet_data.apply(lambda row: row.astype(str).str.contains("Mid#", na=False).any(), axis=1)]
        
        if not mid_cell.empty:  # Proceed only if "Mid#" is found
            mid_text = mid_cell.iloc[0, 0]  # Assuming "Mid#" is in the first column
            mid_value = mid_text.split("Mid#")[-1].strip()  # Extract the value after "Mid#"
            
            # Search for "DBA:" in the sheet
            dba_cell = sheet_data[sheet_data.apply(lambda row: row.astype(str).str.contains("DBA:", na=False).any(), axis=1)]
            if not dba_cell.empty:
                dba_text = dba_cell.iloc[0, 0]  # Assuming "DBA:" is in the first column
                dba_value = dba_text.split("DBA:")[-1].strip()  # Extract the value after "DBA:"
            else:
                # Search for "Merchant :" if "DBA:" is not present
                merchant_cell = sheet_data[sheet_data.apply(lambda row: row.astype(str).str.contains("Merchant :", na=False).any(), axis=1)]
                if not merchant_cell.empty:
                    merchant_text = merchant_cell.iloc[0, 0]  # Assuming "Merchant :" is in the first column
                    dba_value = merchant_text.split("Merchant :")[-1].strip()  # Extract the value after "Merchant :"
                else:
                    dba_value = None  # Default to None if neither is found
            
            # Extract the "volume" value
            volume_value = None
            per_transaction_col = sheet_data.apply(lambda col: col.astype(str).str.contains("Per Transaction", na=False).any(), axis=0)
            if per_transaction_col.any():
                col_index = per_transaction_col[per_transaction_col].index[0]  # Index of the "Per Transaction" column
                # Get the second value underneath the "Per Transaction" header
                volume_row_index = sheet_data.index[sheet_data.iloc[:, col_index].notna()][0] + 2  # Second row under header
                if volume_row_index < len(sheet_data):  # Ensure row exists
                    volume_value = sheet_data.iloc[volume_row_index, col_index]
                    volume_value = round(float(volume_value), 2) if pd.notna(volume_value) else None  # Round to 2 decimals
            
            # Add the default column value
            total_residual = 0.00
            
            # Extract "paydiverse_residual"
            paydiverse_residual_value = None
            total_commissions_row = sheet_data[sheet_data.apply(lambda row: row.astype(str).str.contains("Total Commissions due", na=False).any(), axis=1)]
            if not total_commissions_row.empty:
                row_index = total_commissions_row.index[0]  # Row containing "Total Commissions due"
                col_index = sheet_data.columns[sheet_data.iloc[row_index].notna()][0]  # First non-NaN column
                if col_index + 1 < len(sheet_data.columns):  # Ensure next column exists
                    raw_value = str(sheet_data.iloc[row_index, col_index + 1])
                    # Check for parentheses to determine negativity
                    if raw_value.startswith("(") and raw_value.endswith(")"):
                        paydiverse_residual_value = -float(raw_value.strip("()"))
                    else:
                        paydiverse_residual_value = float(raw_value)
                    paydiverse_residual_value = round(paydiverse_residual_value, 2) if pd.notna(paydiverse_residual_value) else None  # Round to 2 decimals

            # Append all extracted values to the list
            mids_data.append({
                "mid": mid_value,
                "dba": dba_value,
                "volume": volume_value,
                "total_residual": total_residual,
                "paydiverse_residual": paydiverse_residual_value
            })

    # Convert to DataFrame
    mids_df = pd.DataFrame(mids_data)

    # Save or display the DataFrame
    return mids_df  # Or save to a file using mids_df.to_csv("output.csv", index=False)


def process_payscout_file(file_path, filename):
    """
    Process the given file (Excel or CSV), find valid data rows, and return the processed DataFrame.

    Parameters:
        file_path (str): Path to the file to be processed.

    Returns:
        DataFrame: A DataFrame with processed data (for Excel, the first valid sheet is returned).
    """
    mapping = {
        'Merchant Identifier': 'mid',
        'DBA': 'dba',
        'Volume': 'volume',
        'Revenue': 'total_residual',
        'Commission': 'paydiverse_residual'
    }

    def parse_financial_value(value):
        """Convert financial string values to floats, handling parentheses as negative values."""
        if isinstance(value, str) and '(' in value and ')' in value:
            return -float(value.replace("(", "").replace(")", "").replace("$", "").replace(",", ""))
        return float(str(value).replace("$", "").replace(",", ""))

    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        statement_data = pd.read_excel(file_path, sheet_name=None, header=None)
    elif filename.endswith('.csv'):
        statement_data = {"CSV": pd.read_csv(file_path, header=None)}
    else:
        raise ValueError("Unsupported file format. Please provide an Excel or CSV file.")

    for sheet_name, sheet_data in statement_data.items():
        for i, row in sheet_data.iterrows():
            if all(col in row.values for col in mapping.keys()):
                sheet_data.columns = sheet_data.iloc[i].values
                sheet_data = sheet_data.iloc[i + 1:].reset_index(drop=True)
                sheet_data = sheet_data.rename(columns=mapping)

                sheet_data = sheet_data[sheet_data['mid'].apply(lambda x: pd.notna(x) and isinstance(x, str) and x.isdigit())]
                sheet_data['mid'] = sheet_data['mid'].astype(str)
                sheet_data['dba'] = sheet_data['dba'].astype(str)

                for column in ['volume', 'total_residual', 'paydiverse_residual']:
                    sheet_data[column] = sheet_data[column].apply(parse_financial_value)

                return sheet_data[['mid', 'dba', 'volume', 'total_residual', 'paydiverse_residual']]

    raise ValueError("No valid data found in the provided file.")


def process_seamless_file(file_path):
    sheet_name = 0  # First sheet
    sheet_data = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

    # Define the columns we need
    required_columns = {
        "MID": "mid",
        "Merchant Name": "dba",
        "Volume": "volume",
        "Residual": "total_residual",
        "Payout Total": "paydiverse_residual"
    }
    print(sheet_data)
    # Locate the header row containing all required columns
    header_row_index = None
    for i, row in sheet_data.iterrows():
        if all(col in row.values for col in required_columns.keys()):
            header_row_index = i
            break

    if header_row_index is None:
        raise ValueError("Header row with required columns not found.")

    # Load the data with the identified header row
    data = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row_index)

    # Rename columns based on mapping
    data = data.rename(columns=required_columns)

    # Keep only rows with valid MID (non-empty and containing integers)
    data = data[data["mid"].apply(lambda x: isinstance(x, int) or (isinstance(x, str) and x.isdigit()))]

    # Keep only the required columns
    data = data[list(required_columns.values())]

    return data  


def process_micamp_file(file_path, filename):
    """
    Process a MiCamp file and extract relevant data.

    Parameters:
        file_path (str): Path to the Excel or CSV file.

    Returns:
        DataFrame: Processed data containing 'mid', 'dba', 'total_residual', and 'paydiverse_residual'.
    """
    # Mapping of column names
    mapping = {
        'MID': 'mid',
        'Merchant': 'dba',
        'Total Resale Price': 'total_residual',
        'Total Commission': 'paydiverse_residual'
    }

    def is_valid_merchant_id(value):
        """Checks if the merchant ID is valid (non-NaN and consists only of digits)."""
        return pd.notna(value) and isinstance(value, str) and value.isdigit()

    def parse_financial_value(value):
        """Parses financial values, treating values in parentheses as negative."""
        if isinstance(value, str) and '(' in value and ')' in value:
            return -float(value.replace("(", "").replace(")", "").replace("$", "").replace(",", ""))
        return float(str(value).replace("$", "").replace(",", ""))

    def find_data_rows(df):
        """
        Find the header row based on mapping and return the header row index and valid data rows.
        
        Parameters:
            df (DataFrame): The input DataFrame.
        
        Returns:
            DataFrame: Processed DataFrame with relevant columns renamed and filtered.
        """
        header_row = None
        for i, row in df.iterrows():
            # Check if all mapped columns exist in this row
            if all(col in row.values for col in mapping.keys()):
                header_row = i
                break

        if header_row is not None:
            # Set the header row as the columns of the DataFrame
            df.columns = df.iloc[header_row].values
            df = df.iloc[header_row + 1:].reset_index(drop=True)

            # Rename columns based on mapping
            df = df.rename(columns=mapping)

            # Filter rows where 'mid' is valid
            valid_df = df[df['mid'].apply(is_valid_merchant_id)]

            # Convert relevant columns to appropriate types
            valid_df['mid'] = valid_df['mid'].astype(str)
            valid_df['dba'] = valid_df['dba'].astype(str)
            for column in ['total_residual', 'paydiverse_residual']:
                valid_df[column] = valid_df[column].apply(parse_financial_value)

            # Add a 'volume' column filled with 0 before 'total_residual'
            valid_df.insert(valid_df.columns.get_loc('total_residual'), 'volume', 0)

            # Extract only the required columns
            return valid_df[['mid', 'dba', 'volume', 'total_residual', 'paydiverse_residual']]

        return pd.DataFrame()  # Return empty DataFrame if no valid header or data rows

    def load_file(file_path, filename):
        """Load the file and return a dictionary of DataFrames for Excel or a single DataFrame for CSV."""
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            return pd.read_excel(file_path, sheet_name=None, header=None)
        elif file_path.endswith('.csv'):
            return {"CSV": pd.read_csv(file_path, header=None)}
        else:
            raise ValueError("Unsupported file format. Please provide an Excel or CSV file.")

    try:
        # Load the file
        statement_data = load_file(file_path, filename)
        all_processed_data = []

        # Process each sheet in the file
        for sheet_name, sheet_data in statement_data.items():
            processed_data = find_data_rows(sheet_data)
            if not processed_data.empty:
                all_processed_data.append(processed_data)

        # Combine all processed sheets into a single DataFrame
        if all_processed_data:
            return pd.concat(all_processed_data, ignore_index=True)

    except ValueError as e:
        print(f"Error processing file: {e}")


def process_merchant_industry_file(file_path, filename):
    """
    Process a Merchant Industry file and extract relevant data.

    Parameters:
        file_path (str): Path to the Excel or CSV file.

    Returns:
        DataFrame: Processed data containing 'mid', 'dba', 'total_residual', and 'paydiverse_residual'.
    """
    try:
        # Mapping of column names
        if filename.endswith('.csv'):
            df = pd.read_csv(file_path, header=None)
        else:
            df = pd.read_excel(file_path, header=None)

        # Identify the header row dynamically
        header_row_index = df[df.apply(lambda row: row.astype(str).str.contains('Processor', na=False).any(), axis=1)].index[0]

        # Set the header and process the data
        df.columns = df.iloc[header_row_index]  # Use the identified header row
        df = df.iloc[header_row_index + 1:]  # Skip to the data rows
        df = df.reset_index(drop=True)  # Reset index for clean output

        # Optional: Clean up column names (strip whitespace, etc.)
        df.columns = df.columns.str.strip()

        # Keep only the required columns
        columns_to_keep = ['MID', 'Name', 'Volume', 'Income', 'Net']
        df = df[columns_to_keep]

        # Drop rows with NaN values or incomplete data
        df = df.dropna()

        # Convert 'Volume', 'Income', and 'Net' to float
        df['Volume'] = df['Volume'].replace({'\$': '', ',': ''}, regex=True).astype(float)
        df['Income'] = df['Income'].replace({'\$': '', ',': ''}, regex=True).astype(float)
        df['Net'] = df['Net'].replace({'\$': '', ',': ''}, regex=True).astype(float)

        # Rename the headers
        df.columns = ['mid', 'dba', 'volume', 'total_residual', 'paydiverse_residual']

        # Reset index
        df = df.reset_index(drop=True)
        return df

    except ValueError as e:
        print(f"Error processing file: {e}")


def extract_month_from_filename(file_name):
    """Extracts the month and year from the file name."""
    try:
        parts = file_name.split()
        month_str = parts[-2]
        year_str = parts[-1].split('.')[0]
        month_num = datetime.strptime(month_str, "%B").month
        return month_num, int(year_str)
    except Exception as e:
        raise ValueError(f"Failed to extract month and year from file name: {file_name}") from e


def load_file(file_path, filename):
    """Load a file (Excel or CSV) without assuming the first row contains headers."""
    if filename.endswith(('.xlsx', '.xls')):
        return pd.read_excel(file_path, header=None)
    elif filename.endswith('.csv'):
        return pd.read_csv(file_path, header=None)
    else:
        raise ValueError("Unsupported file format. Please provide an Excel or CSV file.")


def find_and_set_headers(df, needed_columns):
    """Find the row where any of the needed columns is present and set it as the header."""
    for i, row in df.iterrows():
        row_values = [str(val).strip() if pd.notna(val) else '' for val in row.values]
        if any(col in row_values for col in needed_columns):
            df.columns = row  # Set the row as columns
            return df[i + 1:].reset_index(drop=True)  # Skip the header row and reset the index
    raise ValueError("Could not find any row with the required columns")


def clean_column_names(df):
    """Cleans column names by removing spaces and special characters."""
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.lower()
        .str.replace(r'[^a-z0-9_]', '', regex=True)
    )


def filter_by_month_and_year(df, column_name, target_month, target_year):
    """Filters the DataFrame for rows matching the target month and year."""
    try:
        df[column_name] = pd.to_datetime(df[column_name], errors='coerce')
        return df[
            (df[column_name].dt.month == target_month) & 
            (df[column_name].dt.year == target_year)
        ]
    except Exception as e:
        raise ValueError(f"Failed to filter by month and year: {str(e)}") from e


def process_ccbill_file(file_path, filename):
    """Process the file and return aggregated data with specified column names."""
    # Column mappings
    header_columns = {
        "Partner ID": "partner_id",
        "Referred Merchant Company Name": "referred_merchant_company_name",
        "Gross Processing Volume": "gross_processing_volume",
        "Total Payout (less delivery)": "total_payout_less_delivery",
        "Commission Payout Date": "commission_payout_date"
    }

    cleaned_column_mapping = {
        'partnerid': 'partner_id',
        'referredmerchantcompanyname': 'referred_merchant_company_name',
        'grossprocessingvolume': 'gross_processing_volume',
        'totalpayoutlessdelivery': 'total_payout_less_delivery',
        'commissionpayoutdate': 'commission_payout_date'
    }

    final_column_mapping = {
        'partner_id': 'mid',
        'referred_merchant_company_name': 'dba',
        'gross_processing_volume': 'volume',
        'total_payout_less_delivery': 'paydiverse_residual'
    }

    # Extract month and year from filename
    target_month, target_year = extract_month_from_filename(filename)

    # Process the file
    df = load_file(file_path, filename)
    df = find_and_set_headers(df, header_columns.keys())  # Set the correct header row
    clean_column_names(df)
    df = df.rename(columns=cleaned_column_mapping)
    
    # Filter and select columns
    filtered_df = filter_by_month_and_year(df, 'commission_payout_date', target_month, target_year)
    df_selected = filtered_df[list(final_column_mapping.keys())]

    # Convert numeric columns
    for col in ['gross_processing_volume', 'total_payout_less_delivery']:
        df_selected[col] = pd.to_numeric(
            df_selected[col].replace('[\$,]', '', regex=True),
            errors='coerce'
        )

    # Aggregate data
    aggregated_df = df_selected.groupby('partner_id').agg({
        'referred_merchant_company_name': 'first',
        'gross_processing_volume': 'sum',
        'total_payout_less_delivery': 'sum'
    }).reset_index()

    # Round numeric columns
    numeric_columns = ['gross_processing_volume', 'total_payout_less_delivery']
    aggregated_df[numeric_columns] = aggregated_df[numeric_columns].round(2)

    # Rename columns to final names
    final_df = aggregated_df.rename(columns=final_column_mapping)
    
    # Add total_residual column with default value 0.00
    final_df.insert(3, 'total_residual', 0.00)
    
    # Reset index
    final_df = final_df.reset_index(drop=True)

    return final_df


def safe_float_convert(value) -> float:
    if value is None:
        return 0.00
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    if isinstance(value, str) and value.startswith('='):
        return 0.00
    try:
        return round(float(value), 2)
    except (ValueError, TypeError):
        return 0.00
    

def find_merchant_blocks(sheet) -> List[Dict]:
    merchant_blocks = []
    HEADER_ROW = 6
    MID_ROW = 5
    DBA_ROW = 4
    
    col = 1
    while col < sheet.max_column:
        mid_value = sheet.cell(row=MID_ROW, column=col).value
        
        if mid_value is not None and str(mid_value).startswith('635400000000'):
            dba_value = sheet.cell(row=DBA_ROW, column=col).value
            
            current_col = col
            count_amount_col = None
            gross_total_col = None
            iso_payout_col = None
            
            for search_col in range(current_col, min(current_col + 8, sheet.max_column + 1)):
                header_value = sheet.cell(row=HEADER_ROW, column=search_col).value
                if header_value:
                    header_str = str(header_value)
                    if 'Merchant COUNT/Amount' in header_str and not count_amount_col:
                        count_amount_col = search_col
                    elif 'Gross Total' in header_str and not gross_total_col:
                        gross_total_col = search_col
                    elif 'ISO Payout' in header_str and not iso_payout_col:
                        iso_payout_col = search_col

            if count_amount_col and gross_total_col and iso_payout_col:
                merchant_block = {
                    'mid': str(mid_value),
                    'dba': str(dba_value),
                    'count_amount_col': count_amount_col,
                    'gross_total_col': gross_total_col,
                    'iso_payout_col': iso_payout_col
                }
                merchant_blocks.append(merchant_block)
            
            col = iso_payout_col + 1 if iso_payout_col else col + 1
        else:
            col += 1
    
    return merchant_blocks


def extract_merchant_data(sheet, merchant_block: Dict) -> Dict:
    results = {
        'mid': merchant_block['mid'],
        'dba': merchant_block['dba'],
        'volume': 0.00,
        'total_residual': 0.00,
        'paydiverse_residual': 0.00
    }
    
    for row in range(6, sheet.max_row + 1):
        description = sheet.cell(row=row, column=1).value
        if description:
            if "Visa/MC/Discover Interchange Dues" in str(description):
                value = sheet.cell(row=row, column=merchant_block['count_amount_col']).value
                results['volume'] += safe_float_convert(value)
            elif "AMEX Sponsorship Fee" in str(description):
                value = sheet.cell(row=row, column=merchant_block['count_amount_col']).value
                results['volume'] += safe_float_convert(value)
    
    total = 0.00
    for row in range(6, sheet.max_row + 1):
        description = sheet.cell(row=row, column=1).value
        if description and "Total Income" in str(description):
            break
        value = sheet.cell(row=row, column=merchant_block['gross_total_col']).value
        total += safe_float_convert(value)
    results['total_residual'] = round(total, 2)
    
    for row in range(1, sheet.max_row + 1):
        description = sheet.cell(row=row, column=1).value
        if description and "NET residuals for payout" in str(description):
            value = sheet.cell(row=row, column=merchant_block['iso_payout_col']).value
            results['paydiverse_residual'] = safe_float_convert(value)
            break
    
    return results


def process_pepper_pay_file(file_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb[wb.sheetnames[1]]
    
    merchant_blocks = find_merchant_blocks(sheet)
    all_results = []
    
    for merchant_block in merchant_blocks:
        merchant_data = extract_merchant_data(sheet, merchant_block)
        all_results.append(merchant_data)
    
    wb.close()
    
    if not all_results:
        return pd.DataFrame()
    
    df = pd.DataFrame(all_results)
    
    numeric_cols = ['volume', 'total_residual', 'paydiverse_residual']
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').round(2)
    
    df['mid'] = df['mid'].astype(str)
    df['dba'] = df['dba'].astype(str)

    return df


def safe_float_convert(value) -> float:
    if value is None:
        return 0.00
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    if isinstance(value, str) and value.startswith('='):
        return 0.00
    try:
        return round(float(value), 2)
    except (ValueError, TypeError):
        return 0.00

def get_column_indices(sheet) -> Dict[str, int]:
    column_mapping = {
        'Merchant ID': 'mid',
        'Merchant Name': 'dba', 
        'P Volume': 'volume',
        'Net Income': 'total_residual',
        'Agent Residual': 'paydiverse_residual'
    }
    
    indices = {}
    header_row = None
    
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value == "Merchant ID":
                header_row = row
                break
        if header_row:
            break
            
    if header_row:
        for col in range(1, sheet.max_column + 1):
            header = str(sheet.cell(row=header_row, column=col).value or "").strip()
            if header in column_mapping:
                indices[column_mapping[header]] = col
                
    return header_row, indices

def extract_quantum_merchant_data(sheet, row: int, col_indices: Dict[str, int]) -> Dict:
    results = {
        'mid': str(sheet.cell(row=row, column=col_indices['mid']).value),
        'dba': str(sheet.cell(row=row, column=col_indices['dba']).value),
        'volume': safe_float_convert(sheet.cell(row=row, column=col_indices['volume']).value),
        'total_residual': safe_float_convert(sheet.cell(row=row, column=col_indices['total_residual']).value),
        'paydiverse_residual': safe_float_convert(sheet.cell(row=row, column=col_indices['paydiverse_residual']).value)
    }
    return results

def extract_all_merchant_data(file_path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    header_row, column_indices = get_column_indices(sheet)
    
    if not header_row or len(column_indices) != 5:
        wb.close()
        return pd.DataFrame()
    
    data = []
    for row in range(header_row + 1, sheet.max_row + 1):
        merchant_data = extract_quantum_merchant_data(sheet, row, column_indices)
        data.append(merchant_data)
    
    wb.close()
    
    if not data:
        return pd.DataFrame()
        
    df = pd.DataFrame(data)
    
    numeric_cols = ['volume', 'total_residual', 'paydiverse_residual']
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce').round(2)
    
    df['mid'] = df['mid'].astype(str)
    df['dba'] = df['dba'].astype(str)
    
    return df


def process_quantum_file(file_path):
    df = extract_all_merchant_data(file_path)
    with pd.option_context('display.float_format', '{:.2f}'.format):
            return df
    

def is_valid_merchant_seamless_id(mid):
   """Check if MID is valid."""
   return pd.notna(mid) and str(mid).replace('-', '').isdigit()


def ceil_to_two_decimals_seamless(value):
   """Ceil a number to two decimal places."""
   if pd.notna(value):
       try:
           return math.ceil(float(value) * 100) / 100
       except (ValueError, TypeError):
           return value
   return value

def process_seamless_paynote_file(file_path):
   required_columns = {
       "MID": "mid",
       "Merchant Name": "dba",
       "Debit Processing Volume": "volume",  
       "Total Partner Income": "total_residual",
       "Total Partner Payout": "paydiverse_residual"
   }

   df = pd.read_excel(file_path, header=None)
   
   header_row = None
   for i, row in df.iterrows():
       if all(col in row.values for col in required_columns.keys()):
           header_row = i
           break
           
   if header_row is None:
       raise ValueError("Required columns not found")
       
   data = pd.read_excel(file_path, header=header_row, dtype={'MID': str})
   data = data.rename(columns=required_columns)
   
   # Removed the volume insertion since we're now mapping it directly
   
   data = data[data['mid'].apply(is_valid_merchant_seamless_id)]
   
   data['mid'] = data['mid'].astype(str).str.replace('.0', '')
   data['dba'] = data['dba'].astype(str)
   
   numeric_cols = ['volume', 'total_residual', 'paydiverse_residual']
   for col in numeric_cols:
       data[col] = data[col].astype(str).replace({r'\$': '', ',': ''}, regex=True)
       data[col] = pd.to_numeric(data[col], errors='coerce')
       data[col] = data[col].apply(ceil_to_two_decimals_seamless).fillna(0.00)

   final_df = data[['mid', 'dba', 'volume', 'total_residual', 'paydiverse_residual']]
   
   pd.set_option('display.max_rows', None)
   pd.set_option('display.float_format', lambda x: '%.2f' % x)
   
   return final_df
